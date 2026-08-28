"""
Excel Report Generator using openpyxl
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ─── Style helpers ────────────────────────────────────────────────────────────

GREEN   = '16a34a'
DARK    = '0f172a'
LIGHT   = 'f8fafc'
BORDER_COLOR = 'e2e8f0'

def _thin_border():
    side = Side(style='thin', color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)

def _header_fill():
    return PatternFill('solid', fgColor=GREEN)

def _alt_fill():
    return PatternFill('solid', fgColor='f1f5f9')

def _write_headers(ws, headers: list[str], row: int = 1):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color='ffffff', size=10)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()

def _write_row(ws, values: list, row: int, alt: bool = False):
    fill = _alt_fill() if alt else None
    for col, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical='center')
        if fill:
            cell.fill = fill

def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def _title_row(ws, title: str, subtitle: str, cols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14, color=DARK)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(size=9, color='64748b')
    ws.row_dimensions[2].height = 16


# ─── Execution Report ─────────────────────────────────────────────────────────

def generate_execution_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Execution Report'
    ws.freeze_panes = 'A6'

    cols = 9
    _title_row(ws, f"Execution Report — {data['period'].capitalize()}",
               f"Period: {data['start']} to {data['end']}", cols)

    # KPI summary row
    kpis = [
        ('Total Signals', data['totalSignals']),
        ('Executed', data['executed']),
        ('Missed', data['missed']),
        ('Wrong Direction', data['wrongDirection']),
        ('Late', data['late']),
        ('Execution Rate', f"{data['executionRate']}%"),
    ]
    for i, (label, value) in enumerate(kpis, start=1):
        ws.cell(row=3, column=i, value=label).font = Font(bold=True, size=8, color='64748b')
        ws.cell(row=4, column=i, value=value).font = Font(bold=True, size=12)

    headers = ['ID', 'Pair', 'Direction', 'Fib Entry', 'Take Profit', 'Stop Loss',
               'Time', 'Session Date', 'Status', 'Exec Rate %', 'Strategy', 'Timeframe']
    _write_headers(ws, headers, row=5)

    for i, sig in enumerate(data['signals']):
        _write_row(ws, [
            sig['id'], sig['pair'], sig['direction'], sig['fib_entry'],
            sig['take_profit'], sig['stop_loss'], str(sig['issued_at']),
            str(sig['session_date']), sig['status'], sig['execution_rate'],
            sig.get('strategy_name', ''), sig.get('timeframe', ''),
        ], row=6+i, alt=i % 2 == 1)

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Attendance Report ────────────────────────────────────────────────────────

def generate_attendance_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'
    ws.freeze_panes = 'A5'

    cols = 8
    _title_row(ws, f"Attendance Report — {data['period'].capitalize()}",
               f"Period: {data['start']} to {data['end']}", cols)

    kpis = [('Present', data['present']), ('Late', data['late']),
            ('Absent', data['absent']), ('Rate', f"{data['attendanceRate']}%")]
    for i, (l, v) in enumerate(kpis, start=1):
        ws.cell(row=3, column=i, value=l).font = Font(bold=True, size=8, color='64748b')
        ws.cell(row=4, column=i, value=v).font = Font(bold=True, size=12)

    headers = ['Trader', 'Email', 'Date', 'Status', 'Check-in', 'GPS Valid',
               'GPS Distance (m)', 'IP Address', 'Device', 'Validated', 'Admin Note']
    _write_headers(ws, headers, row=5)

    for i, r in enumerate(data['records']):
        name = f"{r.get('user__first_name','')} {r.get('user__last_name','')}".strip()
        _write_row(ws, [
            name, r.get('user__email', ''), str(r['date']), r['status'],
            str(r.get('check_in_time', '')), 'Yes' if r.get('gps_valid') else 'No',
            r.get('gps_distance_m', ''), r.get('ip_address', ''),
            r.get('device_info', ''), 'Yes' if r.get('is_validated') else 'No',
            r.get('admin_note', ''),
        ], row=6+i, alt=i % 2 == 1)

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Leaderboard ─────────────────────────────────────────────────────────────

def generate_leaderboard_xlsx(data: list[dict], period: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leaderboard'
    ws.freeze_panes = 'A5'

    start = data[0]['periodStart'] if data else ''
    end   = data[0]['periodEnd']   if data else ''
    cols = 10
    _title_row(ws, f"Discipline Leaderboard — {period.capitalize()}",
               f"Period: {start} to {end} | Ranked by SOP (not profit)", cols)

    ws.cell(row=3, column=1,
            value='⚠ Traders are ranked by SOP compliance metrics. Profit/loss is excluded.').font = Font(size=9, color='dc2626', bold=True)

    headers = ['Rank', 'Trader', 'Email', 'Account', 'Employee ID',
               'Exec Rate %', 'Compliance %', 'Entry Acc %', 'Timing Acc %',
               'Late Entries', 'SOP Score']
    _write_headers(ws, headers, row=4)

    MEDAL = {1: 'fef08a', 2: 'e2e8f0', 3: 'fed7aa'}
    for i, r in enumerate(data):
        row_num = 5 + i
        values = [
            r['rank'], r['name'], r['email'], r.get('accountNumber', ''),
            r.get('employeeId', ''), r['executionRate'], r['complianceScore'],
            r['entryAccuracy'], r['timingAccuracy'], r['lateEntries'], r['sopScore'],
        ]
        _write_row(ws, values, row_num, alt=i % 2 == 1)
        if r['rank'] in MEDAL:
            fill = PatternFill('solid', fgColor=MEDAL[r['rank']])
            ws.cell(row=row_num, column=1).fill = fill

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Compliance Report ────────────────────────────────────────────────────────

def generate_compliance_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Compliance'
    ws.freeze_panes = 'A5'

    cols = 7
    _title_row(ws, f"Compliance Report — {data['period'].capitalize()}",
               f"Period: {data['start']} to {data['end']}", cols)

    kpis = [('Avg Exec Rate', f"{data['avgExecutionRate']}%"),
            ('Avg Compliance', f"{data['avgComplianceScore']}%"),
            ('Traders', len(data['traders']))]
    for i, (l, v) in enumerate(kpis, start=1):
        ws.cell(row=3, column=i, value=l).font = Font(bold=True, size=8, color='64748b')
        ws.cell(row=4, column=i, value=v).font = Font(bold=True, size=12)

    headers = ['Trader', 'Email', 'Account', 'Exec Rate %', 'Compliance %',
               'Entry Acc %', 'Timing Acc %', 'Late Entries']
    _write_headers(ws, headers, row=5)

    for i, t in enumerate(data['traders']):
        _write_row(ws, [
            t['name'], t['email'], t.get('accountNumber', ''),
            t['executionRate'], t['complianceScore'],
            t['entryAccuracy'], t['timingAccuracy'], t['lateEntries'],
        ], row=6+i, alt=i % 2 == 1)

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Session Report ───────────────────────────────────────────────────────────

def generate_session_xlsx(data: dict) -> bytes:
    wb = Workbook()

    # Sheet 1: Execution
    ex = data['execution']
    ws1 = wb.active
    ws1.title = 'Execution'
    _title_row(ws1, 'End-of-Session — Execution', f"Date: {data['sessionDate']}", 9)
    headers1 = ['Pair', 'Direction', 'Fib', 'TP', 'SL', 'Time', 'Status', 'Exec%']
    _write_headers(ws1, headers1, row=3)
    for i, sig in enumerate(ex['signals']):
        _write_row(ws1, [
            sig['pair'], sig['direction'], sig['fib_entry'],
            sig['take_profit'], sig['stop_loss'], str(sig['issued_at']),
            sig['status'], sig['execution_rate'],
        ], row=4+i, alt=i % 2 == 1)
    _auto_width(ws1)

    # Sheet 2: Attendance
    att = data['attendance']
    ws2 = wb.create_sheet('Attendance')
    _title_row(ws2, 'End-of-Session — Attendance', f"Date: {data['sessionDate']}", 6)
    headers2 = ['Trader', 'Status', 'Check-in', 'GPS Valid', 'IP', 'Validated']
    _write_headers(ws2, headers2, row=3)
    for i, r in enumerate(att['records']):
        name = f"{r.get('user__first_name','')} {r.get('user__last_name','')}".strip()
        _write_row(ws2, [
            name, r['status'], str(r.get('check_in_time', '')),
            'Yes' if r.get('gps_valid') else 'No',
            r.get('ip_address', ''), 'Yes' if r.get('is_validated') else 'No',
        ], row=4+i, alt=i % 2 == 1)
    _auto_width(ws2)

    # Sheet 3: Top 5 leaderboard
    ws3 = wb.create_sheet('Leaderboard')
    _title_row(ws3, 'End-of-Session — SOP Leaderboard', f"Date: {data['sessionDate']}", 5)
    headers3 = ['Rank', 'Trader', 'Exec Rate', 'Compliance', 'SOP Score']
    _write_headers(ws3, headers3, row=3)
    for i, r in enumerate(data['leaderboard']):
        _write_row(ws3, [
            f"#{r['rank']}", r['name'],
            f"{r['executionRate']}%", f"{r['complianceScore']}%", r['sopScore'],
        ], row=4+i, alt=i % 2 == 1)
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
